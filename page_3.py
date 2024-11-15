import openai
import os
import openai
import streamlit as st
import requests
from openai import OpenAI
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile


client = OpenAI(api_key='OPENAI_API-KEY') 

st.markdown("Notice inconsistencies with foot traffic data? Tell us & we'll make changes!")
st.sidebar.markdown("Foot Traffic Data Collection ❄️")

q1 = st.text_input('Enter the address you want to provide foot traffic updates for: ')
q2 = st.time_input('What time are you providing input for: ')
day_of_week = st.radio('What day are you providing input for: ', ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])
#traffic_status = st.radio('How crowded is this place right now?', ['Very crowded', 'Somewhat crowded', 'Average', 'Only a few people', 'No one else'])
traffic_status = st.slider(
    'How crowded is this place right now? (0 = very empty, 5 = very crowded)',
    min_value=0,
    max_value=5,
    value=2, 
    format="%d", 
    step=1,  )

openai.api_key = os.environ["OPENAI_API_KEY"]
client = OpenAI()

url = st.text_input('Enter the website link of the destination')

def get_web_content(url):
    # Fetch the webpage
    web_content = ""
    response = requests.get(url)

    # Parse the HTML
    soup = BeautifulSoup(response.text, 'html.parser')

    # Display the title of the webpage
    st.write('Title of the webpage:')
    st.write(soup.title.string)

    # Display the text of the webpage
    st.write('Text of the webpage:')
    article_tags = soup.find_all('article')
    
    for article in article_tags:
        paragraphs = article.find_all('p')
        for p in paragraphs:
            web_content = web_content + str(p.get_text()) + '\n'

    return web_content

# Takes the transcription of the meeting and returns a summary of it via text completions
def abstract_summary_extraction(transcription):
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        temperature=0,
        messages=[
            {
                "role": "system",
                "content": "You are a highly skilled AI trained in language comprehension and summarization. I would like you to read the following text and summarize it into a concise abstract paragraph. Aim to retain the most important points, providing a coherent and readable summary that could help a person understand the main points of the discussion without needing to read the entire text. Please avoid unnecessary details or tangential points."
            },
            {
                "role": "user",
                "content": transcription
            }
        ]
    )
    return response.choices[0].message.content

openai.api_key = os.environ["OPENAI_API_KEY"]
client = OpenAI()

def save_to_excel(data, filename="C:/Users/charv/Downloads/foot_traffic_data.xlsx"):

    os.makedirs(os.path.dirname(filename), exist_ok=True)

    try:
        if os.path.exists(filename):
            try:
                workbook = load_workbook(filename)
                if workbook.sheetnames:
                    with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                        data.to_excel(writer, index=False, sheet_name="FootTraffic", header=False, startrow=writer.sheets["FootTraffic"].max_row)
                else:
                    raise IndexError("At least one sheet must be visible")
            except (InvalidFileException, IndexError):
                with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                    data.to_excel(writer, index=False, sheet_name="FootTraffic")
        else:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                data.to_excel(writer, index=False, sheet_name="FootTraffic")
    
    except Exception as e:
        st.error(f"An error occurred while saving to Excel: {e}")


if st.button('Submit'):
    display_content = get_web_content(url)
   
    st.write(abstract_summary_extraction(display_content))
    st.write('Thank you for providing an update on this location and helping us create better predictions in the future!')
   
    data = pd.DataFrame({
        "Address": [q1],
        "Time": [q2],
        "Day of Week": [day_of_week],
        "Traffic Status": [traffic_status],
        "Website URL": [url]
    })

    save_to_excel(data)
    st.success("Thank you for your feedback! Our team will make changes soon!")  
